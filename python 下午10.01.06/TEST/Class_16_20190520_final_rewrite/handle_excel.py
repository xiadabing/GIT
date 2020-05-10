#如何来封装一些操作
#1、明确需求（读取所有的测试用例，读取一条用例，写入结果）

from openpyxl import load_workbook #可以对已存在的excel进行读写操作
from collections import namedtuple
from TEST.Class_16_20190513_Parser_config.lemon_04_parser_config_4 import do_config


class HandleExecl(object):
    """
    定义处理excel的类
    """
    #config = HandleConfing()   #创建HandleConfing()对象，同时也相当于定义了HandleExecl类属性

    def __init__(self,filename,sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)  #打开文件
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active  #三元运算写法
        # if self.sheetname is None:
        #     self.ws = self.wb.active
        # else:
        #     self.ws = self.wb[self.sheetname]

        self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1,values_only=True))[0]  #创建一个元组类
        self.Cases = namedtuple("Cases", self.sheet_head_tuple)
        self.case_list = []   #定义一个存放所有case命名元组的对象

    def get_cases(self):
        """
        获取所有测试用例
        :return: 存放Case命名元组的列表
        """
        for data in self.ws.iter_rows(min_row=2, values_only=True):
            self.case_list.append(self.Cases(*data))
        return self.case_list

    def get_case(self,row):
        """
        定义获取某一条测试用例
        :param row:  行号
        :return: 返回一个Cases对象
        """
        if isinstance(row,int) and (2 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(mix_row=row, max_row=row,values_only=True))[0]
        else:
            print("传入的行号有误，行号应为大于1的正数")


    def write_result(self,row,actual,result):
        """
        将实际值与测试用例结果保存到excel中
        :param row:  保存到哪一行
        :param actual: 实际值
        :param result: 测试用例执行的结果
        :return:
        """
        if isinstance(row, int) and (2 <= row <= self.ws.max_row):
            self.ws.cell(row=row,column=do_config("excel","actual_col"),value=actual)
            self.ws.cell(row=row, column=do_config("excel","result_col"), value=result)
            self.wb.save(self.filename)


do_excel = HandleExecl(do_config("file path","cases_path"))
if __name__ == '__main__':
    file_name ='test.xlsx'
    one_excel = HandleExecl(filename=file_name)
    #cases = one_excel.get_cases()
    #print(cases)
    one_excel.write_result(3,7,'fail')



