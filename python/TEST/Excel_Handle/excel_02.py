#痛点：通过Worksheet对象中的cell方法，每次调用只能处理一个
#解决：如何解决批量读取

from openpyxl import load_workbook #可以对已存在的excel进行读写操作
from openpyxl import workbook  #可以新建excel文件


#使用load_workbook来实现excel读写操作
#1、打开excel文件(已存在)
wb = load_workbook('test.xlsx')
#一个excel由哪些组成，（文件名、表单、单元格）
#第一个参数为文件名，会返回一个WorkBook对象（相当于整个excel对象）

#2、定位表单
ws = wb['multiply']  #ws为Worksheet对象，（相当于excel中的一个表单）

#3、定位单元格
#方法一，处理指定的某个单元格
# one_cell = ws.cell(row=2,column=2) #one_cell为Cell对象(相当于表单中的一个单元格)
# print(one_cell.value)

#方法二，每遍历一次，处理一个单元格
# Worksheet对象中有如下重要的属性：
#max_row:  单元格最大行
#min_row： 单元格最小行
#max_column：单元格最大列
#min_column：单元格最小列

#Worksheet对象中有如下重要的方法：
#iter_rows: 返回一个生成器，是由每一行的数据构成的元组组成的，
#iter_column: 返回一个生成器，是由每一列的数据构成的元组组成的，
# for row in range(ws.min_row+1,ws.max_row+1):
#     for col_index in range(ws.min_row,ws.max_column+1):
#         data = ws.cell(row=row,column=col_index).value
#         print('值为{}\n类型为：{}\n'.format(data,type(data)))
#如果读取的是数字(int\float),那么读取的数据也是数字类型
#所有的非数字，读取的都是字符串类型

#方法三
# for row_tuple in ws.iter_rows(min_row=2): #每遍历一次，会将某一行的所有单元格对象组成一个元组返回
#     for one_cell in row_tuple:
#         data = one_cell.value
#         print('值为{}\n类型为：{}\n'.format(data, type(data)))


#每遍历一次，会将某一行的所有单元格对象(的值)组成一个元组返回
# for row_tuple in ws.iter_rows(min_row=2,values_only=True):
#     for data in row_tuple:
#         print('值为{}\n类型为：{}\n'.format(data, type(data)))

#方法四、命名元组


pass


