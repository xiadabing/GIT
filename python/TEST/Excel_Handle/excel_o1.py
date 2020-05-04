#将测试数据使用excel来处理
#python中处理excel的模块非常多，比如xlrd、xlwt(读写分开，不太方便)
#使用openpyxl同时支持读写操作、需要暗转猴使用
#pip install openpyxl
#如果报错，将需要将pip更新，python -m pip install -upgrade pip
#注意：openpyxl只能处理后缀名为xlsx的excel文件，其它后缀的文件不能处理

from openpyxl import load_workbook #可以对已存在的excel进行读写操作
from openpyxl import workbook  #可以新建excel文件

#使用load_workbook来实现excel读写操作
#1、打开excel文件(已存在)
wb = load_workbook('test.xlsx')
#一个excel由哪些组成，（文件名、表单、单元格）
#第一个参数为文件名，会返回一个WorkBook对象（相当于整个excel对象）

#2、定位表单
#方法一
ws = wb['multiply']  #ws为Worksheet对象，（相当于excel中的一个表单）
#方法二
# ws = wb.active    #默认获取第一个表单

#3、定位单元格
one_cell = ws.cell(row=2,column=2) #one_cell为Cell对象(相当于表单中的一个单元格)
print(one_cell.value)

#4、读数据不需要关闭，也不需要保存excel
