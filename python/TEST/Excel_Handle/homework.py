from openpyxl import load_workbook #可以对已存在的excel进行读写操作
from openpyxl import workbook  #可以新建excel文件
from collections import namedtuple



wb = load_workbook('test.xlsx')
ws = wb['multiply']
# #获取表头信息
# fields = tuple(ws.iter_rows(max_row=1,values_only=True))[0]
# #定义一个命名元组
# love = namedtuple('love',fields,rename=True)
# case_list = []
# for row_data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
#     #one_data_tuple = row_data
#     #创建一个对象
#     data_tuple = love._make(row_data)
#     case_list.append(data_tuple)
# print(case_list)



# class ExcelHandle:
#     def __init__(self,filename,sheets):
#
#         self.filename = filename
#         self.sheets = sheets
#         self.wb = load_workbook(self.filename)
#         self.ws = self.wb[self.sheets]
#
#     def all_case(self):
#         love = namedtuple('love',tuple(ws.iter_rows(max_row=1,values_only=True))[0],rename=True)
#         case_list = []
#         for row_data in ws.iter_rows(min_row=2,values_only=True):
#             one_data_tuple = row_data
#             data_tuple = love._make(one_data_tuple)
#             case_list.append(data_tuple)
#         return case_list

#方法三
case_list = []
head_namedtuple = namedtuple("head_namedtuple",tuple(ws.iter_rows(max_row=1,values_only=True))[0])
for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
     # case_list.append(head_namedtuple(*data))  #因为data是一个元组，所以需要拆包，类似可变参数
     case_list.append(head_namedtuple._make(data))
print(case_list[0].title)












