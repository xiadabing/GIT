#重构之前的单元测试


import unittest
import inspect
#导入自定义模块
from TEST.lemon_unittest.lemon_unnitest_02 import MathOper
from collections import namedtuple

from openpyxl import load_workbook #可以对已存在的excel进行读写操作

wb = load_workbook('test.xlsx')
ws = wb['multiply']  #ws为Worksheet对象，（相当于excel中的一个表单）

case_list = []
head_namedtuple = namedtuple("head_namedtuple",tuple(ws.iter_rows(max_row=1,values_only=True))[0])
for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
     # case_list.append(head_namedtuple(*data))  #因为data是一个元组，所以需要拆包，类似可变参数
     case_list.append(head_namedtuple._make(data))


def setUpModule():
    """
    打开文件
    :return:
    """
    global file_name
    global file
    print('\n{:=^40s}\n'.format('开始执行测试用例'))
    file_name = 'homework.txt'
    print('打开【{}】文件'.format(file_name))
    file = open(file_name,mode='a',encoding='utf-8')
    file.write('\n{:=^40s}\n'.format('开始执行测试用例'))

def tearDownModule():
    """
    重写父类的方法
    每一个用例执行以后会被调用
    :return:
    """
    print("\n{:=^40s}\n".format('用例执行结束'))
    print('关闭【{}】文件'.format(file_name))
    file.write('{:=^40s}\n'.format('用例执行结束'))
    file.close()
    wb.save('test.xlsx')




class TestMulti(unittest.TestCase):  #继承unittest.TeatCase类
    """
    测试2数相乘
    """
    def test_negatives_multi(self):  #测试用例方法名必须以test开头
        """
        测试2个负数相乘
        :return:
        """
        #能够查看当前运行的实例方法名称
        print('\nRunning Test Method:{}'.format(inspect.stack()[0][3]))
        data_namedtuple = case_list.pop(0)
        case_id = data_namedtuple.case_id
        msg = data_namedtuple.title
        data_1 = data_namedtuple.data_1
        data_2 = data_namedtuple.data_2
        expext_xase = data_namedtuple.expext_xase

        #获取实际结果
        real_result = MathOper(data_1,data_2).multiply()
        #验证实际结果与预期结果是否一致
        expect_result = expext_xase
        #将实际结果写入excel
        ws.cell(row=case_id+1,column=6,value=real_result)
        try:
            self.assertEqual(expect_result,real_result,msg='测试{}失败'.format(msg))

        except AssertionError as e:
            print('具体异常为：{}'.format(e))
            file.write('{},执行结果为：{}\n具体异常为：{}\n'.format(msg, 'fail', e))
            ws.cell(row=case_id + 1, column=7, value='fail')
            raise e
        else:
            file.write('{},执行结果为：{}\n'.format(msg, 'pass'))
            ws.cell(row=case_id + 1, column=7, value='pass')




    def test_neg_pos_multi(self):
        """
        2.测试一个正数一个负数相乘
        :return:
        """
        print('\nRunning Test Method:{}'.format(inspect.stack()[0][3]))
        data_namedtuple = case_list.pop(1)
        case_id = data_namedtuple.case_id
        msg = data_namedtuple.title
        data_1 = data_namedtuple.data_1
        data_2 = data_namedtuple.data_2
        expext_xase = data_namedtuple.expext_xase

        # 获取实际结果
        real_result = MathOper(data_1, data_2).multiply()
        # 验证实际结果与预期结果是否一致
        expect_result = expext_xase
        # 将实际结果写入excel
        ws.cell(row=case_id + 1, column=6, value=real_result)
        try:
            self.assertEqual(expect_result, real_result, msg='测试{}失败'.format(msg))

        except AssertionError as e:
            print('具体异常为：{}'.format(e))
            file.write('{},执行结果为：{}\n具体异常为：{}\n'.format(msg, 'fail', e))
            ws.cell(row=case_id + 1, column=7, value='fail')
            raise e
        else:
            file.write('{},执行结果为：{}\n'.format(msg, 'pass'))
            ws.cell(row=case_id + 1, column=7, value='pass')


    def test_two_zero(self):
            """
            两个0相乘
            :return:
            """
            print('\nRunning Test Method:{}'.format(inspect.stack()[0][3]))
            data_namedtuple = case_list.pop(2)
            case_id = data_namedtuple.case_id
            msg = data_namedtuple.title
            data_1 = data_namedtuple.data_1
            data_2 = data_namedtuple.data_2
            expext_xase = data_namedtuple.expext_xase

            # 获取实际结果
            real_result = MathOper(data_1, data_2).multiply()
            # 验证实际结果与预期结果是否一致
            expect_result = expext_xase
            # 将实际结果写入excel
            ws.cell(row=case_id + 1, column=7, value=real_result)
            try:
                self.assertEqual(expect_result, real_result, msg='测试{}失败'.format(msg))

            except AssertionError as e:
                print('具体异常为：{}'.format(e))
                file.write('{},执行结果为：{}\n具体异常为：{}\n'.format(msg, 'fail', e))
                ws.cell(row=case_id + 1, column=7, value='fail')
                raise e
            else:
                file.write('{},执行结果为：{}\n'.format(msg, 'pass'))
                ws.cell(row=case_id + 1, column=7, value='pass')


    def test_two_pos_muti(self):
            """
            两个正数相乘
            :return:
            """
            print('\nRunning Test Method:{}'.format(inspect.stack()[0][3]))
            data_namedtuple = case_list.pop(3)
            case_id = data_namedtuple.case_id
            msg = data_namedtuple.title
            data_1 = data_namedtuple.data_1
            data_2 = data_namedtuple.data_2
            expext_xase = data_namedtuple.expext_xase

            # 获取实际结果
            real_result = MathOper(data_1, data_2).multiply()
            # 验证实际结果与预期结果是否一致
            expect_result = expext_xase
            # 将实际结果写入excel
            ws.cell(row=case_id + 1, column=6, value=real_result)
            try:
                self.assertEqual(expect_result, real_result, msg='测试{}失败'.format(msg))

            except AssertionError as e:
                print('具体异常为：{}'.format(e))
                file.write('{},执行结果为：{}\n具体异常为：{}\n'.format(msg, 'fail', e))
                ws.cell(row=case_id + 1, column=7, value='fail')
                raise e
            else:
                file.write('{},执行结果为：{}\n'.format(msg, 'pass'))
                ws.cell(row=case_id + 1, column=7, value='pass')


if __name__ == '__main__':
    unittest.main()